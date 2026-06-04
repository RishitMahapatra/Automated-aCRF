"""
pipeline/crf_full_pipeline.py
=============================
Steps 1–5 of the CRF processing pipeline.

Main entry point:
    run_pipeline(pdf_path, session_id) -> dict
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import fitz
import openpyxl
import pandas as pd
from PIL import Image

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import (
    EXCEL_PATH, RENDER_DPI, LINE_MIN_RATIO, TABLE_MARKERS,
    JOIN_KEYS, DOMAIN_SHEETS, RAW_SDTM_SHEET,
    COL_SRC_DATASET, COL_SRC_VARIABLE,
    COL_SDTM_DATASET, COL_SDTM_VAR, COL_SDTM_LABEL,
    get_components_dir, get_annotation_json_path, get_summary_path,
)


# =============================================================================
# UTILITY
# =============================================================================

def strip_domain(raw: str) -> str:
    if not isinstance(raw, str):
        return ""
    return raw.strip().split()[0].upper()


def strip_form_digits(form_code: str) -> str:
    return re.sub(r"\d+$", "", str(form_code or "").strip().upper())


def find_col(columns: list, target: str):
    if target in columns:
        return target

    tl = target.lower().strip()

    for c in columns:
        if str(c).lower().strip() == tl:
            return c

    for c in columns:
        if tl in str(c).lower().strip():
            return c

    return None


# =============================================================================
# FORMAT A LOADER
# =============================================================================

def load_mapping_format_a(xl: pd.ExcelFile) -> dict:
    mapping = {}
    available = xl.sheet_names
    found = [s for s in DOMAIN_SHEETS if s in available]

    if not found:
        return {}

    for sheet in found:
        df = xl.parse(sheet, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        cols = list(df.columns)

        col_src_ds = find_col(cols, COL_SRC_DATASET)
        col_src_var = find_col(cols, COL_SRC_VARIABLE)
        col_sdtm_ds = find_col(cols, COL_SDTM_DATASET)
        col_sdtm_v = find_col(cols, COL_SDTM_VAR)
        col_sdtm_l = find_col(cols, COL_SDTM_LABEL)

        if not col_src_var:
            continue

        for col in [col_src_ds, col_sdtm_ds]:
            if col and col in df.columns:
                df[col] = df[col].ffill()

        df.dropna(subset=[col_src_var], inplace=True)

        for _, row in df.iterrows():
            raw_src_var = str(row.get(col_src_var, "") or "").strip()
            if not raw_src_var or raw_src_var.lower() == "nan":
                continue

            src_tokens = re.split(r"[\n|]+", raw_src_var)
            sdtm_v_raw = str(row.get(col_sdtm_v, "") or "").strip() if col_sdtm_v else ""
            sdtm_tokens = re.split(r"[\n|]+", sdtm_v_raw) if sdtm_v_raw else []

            src_ds = str(row.get(col_src_ds, "") or "").strip() if col_src_ds else ""
            domain = strip_domain(src_ds) or sheet.upper()
            sdtm_ds = str(row.get(col_sdtm_ds, "") or "").strip().upper() if col_sdtm_ds else ""
            sdtm_label = str(row.get(col_sdtm_l, "") or "").strip() if col_sdtm_l else ""

            for i, src_var in enumerate(src_tokens):
                src_var = str(src_var or "").strip().upper()
                if not src_var or src_var == "NAN":
                    continue
                if src_var in JOIN_KEYS:
                    continue

                if sdtm_tokens:
                    sdtm_var = str(sdtm_tokens[i] if i < len(sdtm_tokens) else sdtm_tokens[0]).strip().upper()
                else:
                    sdtm_var = ""

                key = (domain, src_var)
                if key not in mapping:
                    mapping[key] = {
                        "sdtm_dataset": sdtm_ds,
                        "sdtm_variable": sdtm_var,
                        "sdtm_label": sdtm_label,
                    }

    return mapping


# =============================================================================
# FORMAT B LOADER
# =============================================================================

def load_mapping_format_b(excel_path: str) -> dict:
    mapping = {}

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except PermissionError:
        return {}
    except Exception:
        return {}

    if RAW_SDTM_SHEET not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[RAW_SDTM_SHEET]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    for row in rows:
        if not row or len(row) < 11:
            continue

        src_dataset = str(row[1] or "").strip().split("\n")[0].strip().upper()
        sdtm_dataset = str(row[8] or "").strip().split("\n")[0].strip().upper()
        sdtm_var = str(row[9] or "").strip().split("\n")[0].strip().upper()
        sdtm_label = str(row[10] or "").strip().split("\n")[0].strip()

        if not src_dataset or src_dataset in ("NONE", "NAN", ""):
            continue

        raw_var_cell = str(row[2] or "").strip()
        src_variables = [
            v.strip().upper()
            for v in raw_var_cell.split("\n")
            if v and str(v).strip()
        ]

        if not src_variables:
            continue

        for src_var in src_variables:
            if not src_var or src_var in ("NONE", "NAN"):
                continue
            if src_var in JOIN_KEYS:
                continue

            key = (src_dataset, src_var)
            if key not in mapping:
                mapping[key] = {
                    "sdtm_dataset": sdtm_dataset,
                    "sdtm_variable": sdtm_var,
                    "sdtm_label": sdtm_label,
                }

    return mapping


# =============================================================================
# UNIFIED MAPPING LOADER
# =============================================================================

def load_mapping(excel_path=None) -> dict:
    path = Path(excel_path) if excel_path else EXCEL_PATH
    path_str = str(path)

    try:
        xl = pd.ExcelFile(path_str, engine="openpyxl")
        sheet_names = xl.sheet_names
    except PermissionError:
        raise RuntimeError(f"Permission denied: {path}")
    except Exception as e:
        raise RuntimeError(f"Failed to open Excel: {e}")

    has_format_b = RAW_SDTM_SHEET in sheet_names
    has_format_a = any(s in sheet_names for s in DOMAIN_SHEETS)

    if has_format_b:
        mapping = load_mapping_format_b(path_str)
    elif has_format_a:
        mapping = load_mapping_format_a(xl)
    else:
        raise RuntimeError(f"No recognised sheets in Excel. Found: {sheet_names}")

    if not mapping:
        raise RuntimeError("Mapping loaded but is empty.")

    return mapping


def _dataset_summary(mapping: dict) -> dict:
    counts = defaultdict(int)
    for (ds, _raw) in mapping:
        counts[ds] += 1
    return dict(counts)


# =============================================================================
# SDTM RESOLVER
# =============================================================================

def resolve_sdtm(domain: str, raw_var: str, mapping: dict):
    raw_var = str(raw_var or "").strip().upper()
    domain = str(domain or "").strip().upper()

    if not raw_var:
        return None

    result = mapping.get((domain, raw_var))
    if result:
        return result

    for (d, v), val in mapping.items():
        if v == raw_var:
            return val

    return None


# =============================================================================
# HORIZONTAL LINE DETECTION
# =============================================================================

def get_horizontal_lines(page: fitz.Page) -> list:
    page_w = page.rect.width
    min_len = page_w * LINE_MIN_RATIO
    lines_y = []

    for path in page.get_drawings():
        for item in path.get("items", []):
            if item[0] == "l":
                p1, p2 = item[1], item[2]
                if abs(p1.y - p2.y) < 2.0 and abs(p2.x - p1.x) >= min_len:
                    lines_y.append((p1.y + p2.y) / 2.0)

        r = path.get("rect")
        if r and r.height < 3.0 and r.width >= min_len:
            lines_y.append((r.y0 + r.y1) / 2.0)

    lines_y = sorted(lines_y)
    deduped = []
    for y in lines_y:
        if not deduped or y - deduped[-1] > 3.0:
            deduped.append(y)

    return deduped


# =============================================================================
# FOOTER DETECTION
# =============================================================================

def get_footer_y(page: fitz.Page):
    page_h = page.rect.height
    study_id_re = re.compile(r'D\d+C\d+_\w+_V[\d.]+')
    footer_y = None

    for w in page.get_text("words"):
        if study_id_re.search(w[4]) and w[1] > page_h * 0.5:
            if footer_y is None or w[1] < footer_y:
                footer_y = w[1]

    return footer_y


# =============================================================================
# ZONE DETECTION
# =============================================================================

def detect_zones(page: fitz.Page) -> dict:
    page_h = page.rect.height
    page_w = page.rect.width
    lines_y = get_horizontal_lines(page)
    footer_y = get_footer_y(page)

    header_y1 = lines_y[0] if lines_y else page_h * 0.12
    footer_y0 = footer_y if footer_y is not None else page_h * 0.92

    if footer_y0 <= header_y1:
        footer_y0 = header_y1 + 10.0

    body_lines = [y for y in lines_y if header_y1 < y < footer_y0]

    return {
        "page_h": page_h,
        "page_w": page_w,
        "header_y0": 0.0,
        "header_y1": header_y1,
        "body_y0": header_y1,
        "body_y1": footer_y0,
        "footer_y0": footer_y0,
        "footer_y1": page_h,
        "body_lines": body_lines,
        "all_lines": lines_y,
    }


# =============================================================================
# PAGE CLASSIFICATION
# =============================================================================

def classify_page(page: fitz.Page) -> str:
    text = page.get_text()
    if all(marker in text for marker in TABLE_MARKERS):
        return "TABLE"
    return "FORM"


# =============================================================================
# FORM CODE EXTRACTION
# =============================================================================

def extract_form_code(page: fitz.Page):
    text = page.get_text()

    m = re.search(r'Form:\s+.+?\(([A-Z]{1,6}\d*)\)', text, re.IGNORECASE)
    if m:
        return m.group(1).upper()

    m = re.search(r'\b([A-Z]{2,4}\d?)\b', text[:300])
    return m.group(1).upper() if m else None


# =============================================================================
# WORD EXTRACTION
# =============================================================================

def get_component_words(page: fitz.Page, y0: float, y1: float) -> list:
    words = page.get_text("words")
    zone_words = [
        w for w in words
        if y0 < (w[1] + w[3]) / 2.0 < y1
    ]
    return sorted(zone_words, key=lambda w: (round(w[1], 1), w[0]))


def words_to_text(words: list) -> str:
    return " ".join(w[4] for w in words)


# =============================================================================
# COMPONENT CLASSIFICATION
# =============================================================================

def is_version_component(text: str) -> bool:
    tokens = text.strip().split()
    if not tokens:
        return False
    return tokens[0].strip().startswith("(Version:AZ")


def is_table_header_component(text: str) -> bool:
    tokens = text.strip().split()
    if not tokens:
        return False
    return tokens[0].strip().upper() == "FIELD"


# =============================================================================
# NUMBER EXTRACTION
# =============================================================================

_PAT_NUM = re.compile(r'\b(\d{1,3})\b')


def extract_field_number_last(text: str):
    matches = list(_PAT_NUM.finditer(text.strip()))
    if not matches:
        return None, "NOT_FOUND"

    last = matches[-1]
    method = "LAST_SINGLE" if len(matches) == 1 else f"LAST_OF_{len(matches)}"
    return int(last.group(1)), method


def extract_index_and_raw_from_table_words(words: list):
    if not words:
        return None, None

    by_x = sorted(words, key=lambda w: w[0])

    index_number = None
    index_x0 = None

    for w in by_x:
        token = w[4].strip()
        if re.fullmatch(r'\d{1,3}', token):
            index_number = int(token)
            index_x0 = w[0]
            break

    if index_number is None:
        return None, None

    raw_variable = None
    for w in by_x:
        if w[0] <= index_x0:
            continue

        token = w[4].strip().upper()
        if re.fullmatch(r'[A-Z][A-Z0-9_]{1,49}', token):
            raw_variable = token
            break

    return index_number, raw_variable


# =============================================================================
# PAGE RENDER
# =============================================================================

def render_page_to_pil(doc: fitz.Document, page_index: int, dpi: int):
    page = doc[page_index]
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=matrix, alpha=False)
    return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)


# =============================================================================
# PROCESS SINGLE PAGE
# =============================================================================

def process_page(doc, page_index, mapping, page_out_dir, master_records, progress_callback=None):
    pnum = page_index + 1
    page = doc[page_index]
    page_type = classify_page(page)
    zones = detect_zones(page)
    form_code = extract_form_code(page) or f"PAGE{pnum}"
    domain = strip_form_digits(form_code)

    body_y0 = zones["body_y0"]
    body_y1 = zones["body_y1"]
    body_lines = zones["body_lines"]
    boundaries = [body_y0] + body_lines + [body_y1]

    img = render_page_to_pil(doc, page_index, RENDER_DPI)
    scale_y = img.height / zones["page_h"]

    if progress_callback:
        progress_callback(pnum, len(doc), page_type, form_code, len(boundaries) - 1)

    for i in range(len(boundaries) - 1):
        comp_y0 = boundaries[i]
        comp_y1 = boundaries[i + 1]
        label = f"component_{i + 1:02d}"

        words = get_component_words(page, comp_y0, comp_y1)
        comp_text = words_to_text(words)

        if not comp_text.strip():
            continue

        if is_version_component(comp_text):
            continue

        if page_type == "TABLE" and is_table_header_component(comp_text):
            continue

        y0_px = max(0, int(comp_y0 * scale_y))
        y1_px = min(img.height, int(comp_y1 * scale_y))

        if y1_px > y0_px:
            img.crop((0, y0_px, img.width, y1_px)).save(
                page_out_dir / f"{label}.png"
            )

        if page_type == "TABLE":
            number, raw_var = extract_index_and_raw_from_table_words(words)
            num_method = "TABLE_LEFTMOST_X"
            raw_method = "TABLE_DIRECT" if raw_var else "TABLE_NO_RAW"
            sdtm_info = resolve_sdtm(domain, raw_var, mapping) if raw_var else None
        else:
            number, num_method = extract_field_number_last(comp_text)
            raw_var = None
            raw_method = "PENDING_TABLE_LINK"
            sdtm_info = None

        record = {
            "page": pnum,
            "page_type": page_type,
            "form_code": form_code,
            "domain": domain,
            "component": label,
            "y0_pts": round(comp_y0, 2),
            "y1_pts": round(comp_y1, 2),
            "y0_px": y0_px,
            "y1_px": y1_px,
            "text_preview": comp_text[:80],
            "field_number": number,
            "num_method": num_method,
            "raw_variable": raw_var,
            "raw_method": raw_method,
            "sdtm_dataset": sdtm_info["sdtm_dataset"] if sdtm_info else None,
            "sdtm_variable": sdtm_info["sdtm_variable"] if sdtm_info else None,
            "sdtm_label": sdtm_info["sdtm_label"] if sdtm_info else None,
        }
        master_records.append(record)


# =============================================================================
# POST-PROCESSING
# =============================================================================

def link_form_to_table(master_records, mapping):
    table_index = {}

    for rec in master_records:
        if rec["page_type"] == "TABLE" and rec["raw_variable"]:
            fc = rec["form_code"]
            num = rec["field_number"]
            raw = rec["raw_variable"]

            if fc and num is not None:
                table_index.setdefault(fc, {})[num] = raw

    linked = 0
    unlinked = 0

    for rec in master_records:
        if rec["raw_method"] != "PENDING_TABLE_LINK":
            continue

        form_code = rec["form_code"]
        number = rec["field_number"]
        domain = rec["domain"]

        if number is None:
            rec["raw_method"] = "UNRESOLVED"
            unlinked += 1
            continue

        raw_var = table_index.get(form_code, {}).get(number)

        if not raw_var:
            for fc, num_map in table_index.items():
                if strip_form_digits(fc) == domain:
                    raw_var = num_map.get(number)
                    if raw_var:
                        break

        if raw_var:
            sdtm_info = resolve_sdtm(domain, raw_var, mapping)
            rec["raw_variable"] = raw_var
            rec["raw_method"] = "TABLE_LINK"
            rec["sdtm_dataset"] = sdtm_info["sdtm_dataset"] if sdtm_info else None
            rec["sdtm_variable"] = sdtm_info["sdtm_variable"] if sdtm_info else None
            rec["sdtm_label"] = sdtm_info["sdtm_label"] if sdtm_info else None
            linked += 1
        else:
            rec["raw_method"] = "UNRESOLVED"
            unlinked += 1

    return linked, unlinked


# =============================================================================
# APPLY USER CORRECTIONS
# =============================================================================

def apply_corrections(master_records, corrections, mapping):
    correction_map = {
        (c["page"], c["component"]): c
        for c in corrections
        if c.get("sdtm_dataset") and c.get("sdtm_variable")
    }

    applied = 0
    for rec in master_records:
        key = (rec["page"], rec["component"])
        if key in correction_map:
            corr = correction_map[key]
            rec["sdtm_dataset"] = str(corr["sdtm_dataset"]).strip().upper()
            rec["sdtm_variable"] = str(corr["sdtm_variable"]).strip().upper()
            rec["sdtm_label"] = str(corr.get("sdtm_label", "") or "").strip()
            rec["raw_method"] = "USER_CORRECTED"
            applied += 1

    return applied


# =============================================================================
# RUN PIPELINE
# =============================================================================

def run_pipeline(pdf_path, session_id, corrections=None, progress_callback=None):
    pdf_path = Path(pdf_path)
    components_dir = get_components_dir(session_id)
    json_path = get_annotation_json_path(session_id)
    summary_path = get_summary_path(session_id)

    mapping = load_mapping()

    doc = fitz.open(str(pdf_path))
    total_pages = len(doc)

    master_records = []

    for page_index in range(total_pages):
        pnum = page_index + 1
        page_out_dir = components_dir / f"page_{pnum:03d}"
        page_out_dir.mkdir(parents=True, exist_ok=True)

        process_page(
            doc,
            page_index,
            mapping,
            page_out_dir,
            master_records,
            progress_callback,
        )

    doc.close()

    linked, unlinked = link_form_to_table(master_records, mapping)

    if corrections:
        apply_corrections(master_records, corrections, mapping)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(master_records, f, indent=2, ensure_ascii=False)

    form_records = [r for r in master_records if r["page_type"] == "FORM"]
    resolved = sum(1 for r in form_records if r.get("sdtm_variable"))
    unresolved = sum(1 for r in form_records if not r.get("sdtm_variable"))

    summary_lines = [
        "CRF FULL PIPELINE SUMMARY",
        "=" * 72,
        f"PDF    : {pdf_path}",
        f"Pages  : {total_pages}",
        f"Records: {len(master_records)}",
        f"FORM records : {len(form_records)}",
        f"Resolved     : {resolved}",
        f"Unresolved   : {unresolved}",
        f"Linked       : {linked}",
        f"Unlinked     : {unlinked}",
    ]
    summary_path.write_text("\n".join(summary_lines), encoding="utf-8")

    return {
        "records": master_records,
        "resolved": resolved,
        "unresolved": unresolved,
        "total": len(form_records),
        "linked": linked,
        "json_path": json_path,
        "summary_path": summary_path,
        "mapping_stats": _dataset_summary(mapping),
    }