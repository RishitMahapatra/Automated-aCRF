"""
api.py
======
Bridge between JS frontend and Python backend.
Every public method is callable from JS via window.pywebview.api.method_name()
"""

from __future__ import annotations

import csv
import io
import json
import shutil
import traceback
import uuid
import webbrowser
from pathlib import Path

import openpyxl
import webview

from config import ROOT_DIR, OUTPUTS_DIR, EXCEL_PATH, get_session_output_dir

MAPPING_DB_PATH = ROOT_DIR / "assets" / "mapping_database.json"
from bridge.pipeline_bridge import PipelineBridge, run_full_pipeline
from bridge import annotation_bridge
from bridge import editor_state_bridge
from bridge import session_bridge
from bridge.export_bridge import ExportBridge


class Api:
    """Exposed to JS as window.pywebview.api"""

    def __init__(self):
        self._window = None
        self._pdf_path: Path | None = None
        self._session_id: str = ""
        self._acrf_path: Path | None = None
        self._pipeline = PipelineBridge()
        self._export = ExportBridge()
        self._is_dirty: bool = False  # mirrored from JS; read by on_closing without evaluate_js

    # ==========================================================================
    # FILE UPLOAD
    # ==========================================================================

    def select_pdf(self):
        """
        Open file picker and store selected PDF internally.
        """
        try:
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("PDF Files (*.pdf)",),
            )

            if result and len(result) > 0:
                path = Path(result[0])
                if path.exists() and path.suffix.lower() == ".pdf":
                    self._pdf_path = path
                    base = path.stem.strip().replace(" ", "_")
                    suffix = uuid.uuid4().hex[:8]
                    self._session_id = f"{base}_{suffix}"
                    self._export.set_pdf(path)

                    return {
                        "ok": True,
                        "filename": path.name,
                        "path": str(path),
                        "session_id": self._session_id,
                    }

            return {"ok": False, "error": "No file selected"}

        except Exception as e:
            return {"ok": False, "error": str(e)}

    def open_pdf_dialog(self):
        """
        Backward-compatible alias for older frontend code.
        """
        return self.select_pdf()

    # ==========================================================================
    # SESSION
    # ==========================================================================

    def set_session_id(self, session_id):
        self._session_id = str(session_id or "").strip().replace(" ", "_")
        return {"ok": True, "session_id": self._session_id}

    def get_state(self):
        return {
            "ok": True,
            "pdf_loaded": self._pdf_path is not None,
            "pdf_name": self._pdf_path.name if self._pdf_path else None,
            "pdf_path": str(self._pdf_path) if self._pdf_path else None,
            "session_id": self._session_id,
            "page_count": self._export.get_page_count(),
        }

    def restart_session(self):
        """
        Clear current in-memory session state so the frontend can reset the app
        and start a fresh session.
        """
        try:
            self._pdf_path = None
            self._session_id = ""
            self._acrf_path = None
            self._is_dirty = False
            self._export = ExportBridge()
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # PIPELINE
    # ==========================================================================

    def run_pipeline(self, pdf_path=None, session_id=None):
        """
        Run pipeline using either:
        - stored self._pdf_path / self._session_id
        - or optional args from older frontend code
        """
        try:
            if pdf_path:
                self._pdf_path = Path(pdf_path)

            if session_id:
                self._session_id = str(session_id).strip().replace(" ", "_")

            if not self._pdf_path:
                return {"ok": False, "error": "No PDF loaded"}

            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            result = run_full_pipeline(
                pdf_path=self._pdf_path,
                session_id=self._session_id,
            )

            if result.get("ok"):
                self._export.set_pdf(self._pdf_path)

            return result

        except Exception as e:
            return {
                "ok": False,
                "error": str(e),
                "trace": traceback.format_exc(),
            }

    # ==========================================================================
    # PAGE RENDERING
    # ==========================================================================

    def get_page_image(self, page_number, dpi=150):
        try:
            return self._export.render_page_base64(int(page_number), int(dpi))
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_page_count(self):
        try:
            return {"ok": True, "count": self._export.get_page_count()}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # ANNOTATIONS
    # ==========================================================================

    def get_annotations(self):
        try:
            records = annotation_bridge.get_annotations(self._session_id)
            return {"ok": True, "records": records}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_page_annotations(self, page_number):
        try:
            records = annotation_bridge.get_page_annotations(
                self._session_id,
                int(page_number),
            )
            return {"ok": True, "records": records}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_annotation(self, annotation_id):
        try:
            rec = annotation_bridge.get_annotation(self._session_id, annotation_id)
            if rec:
                return {"ok": True, "record": rec}
            return {"ok": False, "error": "Not found"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_suggestions(self, annotation_id):
        try:
            suggestions = annotation_bridge.get_suggestions(
                self._session_id,
                annotation_id,
            )
            return {"ok": True, "suggestions": suggestions}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def update_annotation(
        self,
        annotation_id,
        status,
        sdtm_dataset="",
        sdtm_variable="",
        sdtm_label="",
    ):
        """
        Generic annotation update method for frontend compatibility.
        """
        try:
            return annotation_bridge.update_annotation(
                self._session_id,
                annotation_id,
                status=status,
                sdtm_dataset=(sdtm_dataset or "").strip().upper(),
                sdtm_variable=(sdtm_variable or "").strip().upper(),
                sdtm_label=(sdtm_label or "").strip(),
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def confirm_mapping(self, annotation_id, dataset, variable, label=""):
        try:
            return self.update_annotation(
                annotation_id=annotation_id,
                status="USER_CORRECTED",
                sdtm_dataset=dataset,
                sdtm_variable=variable,
                sdtm_label=label,
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def mark_not_submitted(self, annotation_id):
        try:
            return self.update_annotation(
                annotation_id=annotation_id,
                status="NOT_SUBMITTED",
                sdtm_dataset="",
                sdtm_variable="",
                sdtm_label="Not Submitted",
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def clear_mapping(self, annotation_id):
        try:
            return self.update_annotation(
                annotation_id=annotation_id,
                status="UNMAPPED",
                sdtm_dataset="",
                sdtm_variable="",
                sdtm_label="",
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def remove_annotation(self, annotation_id):
        try:
            return self.update_annotation(
                annotation_id=annotation_id,
                status="REMOVED",
                sdtm_dataset="",
                sdtm_variable="",
                sdtm_label="",
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_annotation(self, annotation_id):
        try:
            rec = annotation_bridge.get_annotation(self._session_id, str(annotation_id or "").strip())
            if rec is None:
                return {"ok": False, "error": "Not found"}
            return {"ok": True, "record": rec}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def update_comment(self, annotation_id, comment=""):
        try:
            return annotation_bridge.update_comment(
                self._session_id,
                annotation_id,
                comment=str(comment or "").strip(),
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # EDITOR STATE SNAPSHOT
    # ==========================================================================

    def save_editor_state(self, state):
        try:
            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            return editor_state_bridge.save_editor_state(
                self._session_id,
                state or {},
            )
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def load_editor_state(self):
        try:
            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            return editor_state_bridge.load_editor_state(self._session_id)
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # STATS
    # ==========================================================================

    def get_stats(self):
        try:
            stats = annotation_bridge.get_stats(self._session_id)
            return {"ok": True, **stats}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # EXPORT
    # ==========================================================================

    def export_pdf(self):
        """
        Existing export path. Kept for backward compatibility.
        """
        try:
            if not self._pdf_path:
                return {"ok": False, "error": "No PDF loaded"}

            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            suggested_name = f"{self._session_id}.pdf"

            save_result = self._window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename=suggested_name,
                file_types=("PDF Files (*.pdf)",),
            )

            if not save_result:
                return {"ok": False, "error": "Export cancelled"}

            if isinstance(save_result, (list, tuple)):
                out_path = Path(save_result[0])
            else:
                out_path = Path(save_result)

            if out_path.suffix.lower() != ".pdf":
                out_path = out_path.with_suffix(".pdf")

            export_result = self._export.export_annotated_pdf(
                pdf_path=self._pdf_path,
                session_id=self._session_id,
                annotations=None,
            )

            if not export_result.get("ok"):
                return export_result

            generated_path = Path(export_result["path"])
            if not generated_path.exists():
                return {"ok": False, "error": "Generated annotated PDF not found"}

            out_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(generated_path, out_path)

            return {
                "ok": True,
                "path": str(out_path),
            }

        except Exception as e:
            return {"ok": False, "error": str(e)}

    def export_pdf_from_images(self, page_images):
        """
        New screenshot-style export path.
        Frontend sends one rendered page image per PDF page.
        """
        try:
            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            if not page_images or not isinstance(page_images, list):
                return {"ok": False, "error": "No page images provided"}

            suggested_name = f"{self._session_id}.pdf"

            save_result = self._window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename=suggested_name,
                file_types=("PDF Files (*.pdf)",),
            )

            if not save_result:
                return {"ok": False, "error": "Export cancelled"}

            if isinstance(save_result, (list, tuple)):
                out_path = Path(save_result[0])
            else:
                out_path = Path(save_result)

            if out_path.suffix.lower() != ".pdf":
                out_path = out_path.with_suffix(".pdf")

            temp_out = (
                get_session_output_dir(self._session_id)
                / f"{self._session_id}_screenshot_export.pdf"
            )

            result = self._export.export_from_page_images(page_images, temp_out)
            if not result.get("ok"):
                return result

            if not temp_out.exists():
                return {"ok": False, "error": "Temporary export PDF not created"}

            out_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(temp_out, out_path)

            return {"ok": True, "path": str(out_path)}

        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # COLOUR
    # ==========================================================================

    def get_dataset_colours(self):
        try:
            colours = annotation_bridge.get_dataset_colours(self._session_id)
            return {"ok": True, "colours": colours}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def set_dataset_colour(self, dataset, colour_key):
        try:
            records = annotation_bridge.get_annotations(self._session_id)

            form_code = ""
            dataset_upper = (dataset or "").strip().upper()

            for rec in records:
                if (rec.get("sdtm_dataset") or "").strip().upper() == dataset_upper:
                    form_code = rec.get("form_code", "")
                    break

            result = annotation_bridge.update_dataset_colour(
                self._session_id,
                form_code,
                dataset_upper,
                colour_key,
            )
            return result

        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # SESSION FILE (.acrf) — Save / Open / Save As
    # ==========================================================================

    def save_session_file(self, editor_state=None, frontend_annotations=None):
        """
        Save to the current .acrf path (or prompt Save As if no path yet).
        frontend_annotations: list of annotation records from the JS Store,
        including user-created ones that only exist in the frontend.
        """
        try:
            if not self._pdf_path:
                return {"ok": False, "error": "No PDF loaded"}
            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            self._merge_frontend_annotations(frontend_annotations)

            if not self._acrf_path:
                return self.save_session_file_as(editor_state, frontend_annotations)

            result = session_bridge.save_session(
                save_path=self._acrf_path,
                session_id=self._session_id,
                pdf_path=self._pdf_path,
                editor_state=editor_state,
            )
            return result

        except Exception as e:
            return {"ok": False, "error": str(e)}

    def save_session_file_as(self, editor_state=None, frontend_annotations=None):
        """
        Prompt for location and save the .acrf session file.
        """
        try:
            if not self._pdf_path:
                return {"ok": False, "error": "No PDF loaded"}
            if not self._session_id:
                return {"ok": False, "error": "No session ID"}

            self._merge_frontend_annotations(frontend_annotations)

            suggested_name = f"{self._session_id}.acrf"

            save_result = self._window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename=suggested_name,
                file_types=("aCRF Session (*.acrf)",),
            )

            if not save_result:
                return {"ok": False, "error": "Save cancelled"}

            if isinstance(save_result, (list, tuple)):
                out_path = Path(save_result[0])
            else:
                out_path = Path(save_result)

            if out_path.suffix.lower() != ".acrf":
                out_path = out_path.with_suffix(".acrf")

            result = session_bridge.save_session(
                save_path=out_path,
                session_id=self._session_id,
                pdf_path=self._pdf_path,
                editor_state=editor_state,
            )

            if result.get("ok"):
                self._acrf_path = out_path

            return result

        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _merge_frontend_annotations(self, frontend_annotations):
        """
        Merge frontend-only annotations (user-drawn, with user_ or
        userdschip_ prefixed IDs) into the backend annotation_data.json
        so they survive save/restore.
        """
        if not frontend_annotations or not isinstance(frontend_annotations, list):
            return
        if not self._session_id:
            return

        from config import get_annotation_json_path
        import json

        json_path = get_annotation_json_path(self._session_id)

        existing = []
        if json_path.exists():
            try:
                existing = json.loads(json_path.read_text(encoding="utf-8"))
                if not isinstance(existing, list):
                    existing = []
            except Exception:
                existing = []

        existing_ids = {
            str(r.get("annotation_id", "")) for r in existing if r
        }

        added = 0
        for rec in frontend_annotations:
            if not rec or not isinstance(rec, dict):
                continue
            ann_id = str(rec.get("annotation_id", ""))
            if not ann_id:
                continue

            if ann_id in existing_ids:
                for i, ex in enumerate(existing):
                    if str(ex.get("annotation_id", "")) == ann_id:
                        # Never overwrite comment — it's always authoritative in
                        # annotation_data.json (saved directly via update_comment)
                        rec_for_merge = {k: v for k, v in rec.items() if k != "comment"}
                        existing[i] = {**ex, **rec_for_merge}
                        break
            else:
                existing.append(rec)
                existing_ids.add(ann_id)
                added += 1

        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(
            json.dumps(existing, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )

    def open_session_file(self):
        """
        Open file picker for .acrf files and load the session.
        """
        try:
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("aCRF Session (*.acrf)",),
            )

            if not result or len(result) == 0:
                return {"ok": False, "error": "No file selected"}

            acrf_path = Path(result[0])
            load_result = session_bridge.open_session(acrf_path)

            if not load_result.get("ok"):
                return load_result

            self._session_id = load_result["session_id"]
            self._pdf_path = Path(load_result["pdf_path"])
            self._acrf_path = acrf_path
            self._export.set_pdf(self._pdf_path)

            return load_result

        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_recent_sessions(self):
        """
        Return the 10 most recently opened .acrf sessions.
        """
        try:
            sessions = session_bridge.get_recent_sessions(limit=10)
            return {"ok": True, "sessions": sessions}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ==========================================================================
    # WINDOW LIFECYCLE
    # ==========================================================================

    def set_dirty(self, dirty):
        """
        JS calls this whenever the dirty state changes so Python knows it
        without having to call evaluate_js inside the closing event handler
        (which would deadlock on the UI thread).
        """
        self._is_dirty = bool(dirty)

    def confirm_close(self):
        """
        Called from JS after the user confirms they want to close the window
        (via the unsaved-changes dialog).  Destroys the window unconditionally.

        _is_dirty must be cleared BEFORE destroy(), because destroy() re-fires
        the closing event — if _is_dirty is still True, on_closing will cancel
        the destroy and re-show the dialog, causing an infinite loop.
        """
        try:
            self._is_dirty = False
            if self._window:
                self._window.destroy()
        except Exception as e:
            print(f"[api] confirm_close error: {e}")

    def open_url(self, url):
        """Open an external URL in the user's default browser."""
        try:
            webbrowser.open(url)
        except Exception as e:
            print(f"[api] open_url error: {e}")

    # ==========================================================================
    # MAPPING DATABASE
    # ==========================================================================

    def select_excel_for_import(self):
        """File picker for Excel files to import into mapping DB."""
        try:
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=('Excel Files (*.xlsx)',),
            )
            if result and len(result) > 0:
                return {"ok": True, "path": str(result[0])}
            return {"ok": False, "error": "No file selected"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def read_excel_preview(self, file_path, header_row=1):
        """Read first 10 rows and all column headers from the Excel file."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i > header_row + 10:
                    break
                cells = [str(c) if c is not None else "" for c in row]
                rows.append({"row_num": i, "cells": cells})
            col_count = ws.max_column or 0
            wb.close()
            return {"ok": True, "rows": rows, "col_count": col_count}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    @staticmethod
    def _col_to_idx(val):
        """Convert Excel column letter (A, B, AA) or 1-based number string to 0-based index.
        Returns -1 for empty/zero/invalid so callers can treat it as 'skip'."""
        if val is None:
            return -1
        s = str(val).strip().upper()
        if not s or s in ("0", ""):
            return -1
        if s.isalpha():
            try:
                # openpyxl converts A→1, Z→26, AA→27, etc.
                return openpyxl.utils.column_index_from_string(s) - 1
            except Exception:
                return -1
        try:
            n = int(s)
            return n - 1 if n > 0 else -1
        except (ValueError, TypeError):
            return -1

    def import_excel_mapping(self, file_path, config):
        """
        Parse an Excel file using the user's column configuration.
        config columns accept Excel letters (A, B, C) or 1-based numbers.
        """
        try:
            path = Path(file_path).expanduser()
            if not path.exists():
                return {
                    "ok": False,
                    "error": (
                        f"File not found: {path}\n"
                        "If the file is on OneDrive, make sure it is fully downloaded "
                        "(right-click → Always keep on this device) before importing."
                    ),
                }
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            entries = []
            src_ds_idx  = self._col_to_idx(config.get("src_dataset_col"))
            raw_var_idx = self._col_to_idx(config.get("raw_variable_col"))
            sdtm_ds_idx = self._col_to_idx(config.get("sdtm_dataset_col"))
            sdtm_v_idx  = self._col_to_idx(config.get("sdtm_variable_col"))
            sdtm_l_idx  = self._col_to_idx(config.get("sdtm_label_col"))
            raw_l_idx   = self._col_to_idx(config.get("raw_label_col"))
            data_start  = int(config.get("data_start_row", 2))

            def _cell(row, idx):
                if idx < 0 or idx >= len(row):
                    return ""
                return str(row[idx] or "").strip()

            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i < data_start:
                    continue
                cells = list(row)
                src_ds   = _cell(cells, src_ds_idx).upper()
                raw_var  = _cell(cells, raw_var_idx).upper()
                sdtm_ds  = _cell(cells, sdtm_ds_idx).upper()
                sdtm_var = _cell(cells, sdtm_v_idx).upper()
                sdtm_lbl = _cell(cells, sdtm_l_idx) if sdtm_l_idx >= 0 else ""
                raw_lbl  = _cell(cells, raw_l_idx) if raw_l_idx >= 0 else ""

                if not raw_var or raw_var in ("NONE", "NAN", ""):
                    continue

                entries.append({
                    "id": str(uuid.uuid4())[:8],
                    "src_dataset": src_ds,
                    "raw_variable": raw_var,
                    "raw_label": raw_lbl,
                    "sdtm_dataset": sdtm_ds,
                    "sdtm_variable": sdtm_var,
                    "sdtm_label": sdtm_lbl,
                    "source": "import",
                })
            wb.close()
            return {"ok": True, "entries": entries, "count": len(entries)}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def save_mapping_db(self, data):
        """Save mapping database JSON to assets/mapping_database.json."""
        try:
            MAPPING_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(MAPPING_DB_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return {"ok": True, "path": str(MAPPING_DB_PATH)}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def load_mapping_db(self):
        """Load mapping database JSON from assets/mapping_database.json."""
        try:
            if not MAPPING_DB_PATH.exists():
                return {"ok": True, "data": None}
            with open(MAPPING_DB_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {"ok": True, "data": data}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def export_mapping_excel(self, entries):
        """Export mapping entries to an Excel file (user picks save location)."""
        try:
            result = self._window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename="mapping_export.xlsx",
                file_types=('Excel Files (*.xlsx)',),
            )
            if not result:
                return {"ok": False, "error": "No location selected"}
            save_path = str(result) if isinstance(result, str) else str(result[0]) if result else None
            if not save_path:
                return {"ok": False, "error": "No location selected"}

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mapping Database"
            headers = ["Source Dataset", "Raw Variable", "Raw Label",
                       "SDTM Dataset", "SDTM Variable", "SDTM Label"]
            ws.append(headers)
            for e in entries:
                ws.append([
                    e.get("src_dataset", ""),
                    e.get("raw_variable", ""),
                    e.get("raw_label", ""),
                    e.get("sdtm_dataset", ""),
                    e.get("sdtm_variable", ""),
                    e.get("sdtm_label", ""),
                ])
            wb.save(save_path)
            return {"ok": True, "path": save_path}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def export_mapping_csv(self, entries):
        """Export mapping entries to a CSV file."""
        try:
            result = self._window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename="mapping_export.csv",
                file_types=('CSV Files (*.csv)',),
            )
            if not result:
                return {"ok": False, "error": "No location selected"}
            save_path = str(result) if isinstance(result, str) else str(result[0]) if result else None
            if not save_path:
                return {"ok": False, "error": "No location selected"}

            with open(save_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Source Dataset", "Raw Variable", "Raw Label",
                                 "SDTM Dataset", "SDTM Variable", "SDTM Label"])
                for e in entries:
                    writer.writerow([
                        e.get("src_dataset", ""),
                        e.get("raw_variable", ""),
                        e.get("raw_label", ""),
                        e.get("sdtm_dataset", ""),
                        e.get("sdtm_variable", ""),
                        e.get("sdtm_label", ""),
                    ])
            return {"ok": True, "path": save_path}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def save_mapping_file(self, data):
        """Save mapping database as a .mtbl file (portable versioned snapshot)."""
        try:
            result = self._window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename="mapping_table.mtbl",
                file_types=('Mapping Table (*.mtbl)',),
            )
            if not result:
                return {"ok": False, "error": "No location selected"}
            save_path = str(result) if isinstance(result, str) else str(result[0]) if result else None
            if not save_path:
                return {"ok": False, "error": "No location selected"}
            if not save_path.endswith(".mtbl"):
                save_path += ".mtbl"

            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return {"ok": True, "path": save_path}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def load_mapping_file(self):
        """Load a .mtbl file."""
        try:
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=('Mapping Table (*.mtbl)',),
            )
            if not result or len(result) == 0:
                return {"ok": False, "error": "No file selected"}
            file_path = str(result[0])
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {"ok": True, "data": data, "path": file_path}
        except Exception as e:
            return {"ok": False, "error": str(e)}