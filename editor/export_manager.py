"""
editor/export_manager.py
=========================
Module 7 — Export Manager

Handles all export operations after annotation review:

  1. Annotated PDF — re-runs crf_annotator.py with updated
     records from the annotation store, including:
       - All USER_CORRECTED mappings
       - NOT_SUBMITTED boxes (grey)
       - Updated colour registry
       - Contents/TOC page with GO TO links for unmapped

  2. Updated annotation_data.json — store state serialised
     with all corrections baked in

  3. Excel mapping summary — RAW → SDTM table for all
     resolved annotations

  4. Feedback to RAW_SDTM_Mappings.xlsx — new mappings
     discovered during review fed back into master mapping
     file so future pipeline runs auto-resolve them

Dependencies:
  editor/annotation_store.py
  pipeline/crf_annotator.py
  openpyxl, pandas
"""

from __future__ import annotations

import json
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict
from typing import Optional

import fitz
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

from editor.annotation_store import AnnotationStore, AnnotationRecord
from editor.canvas_overlay   import (
    DatasetColourRegistry,
    WONG_PALETTE,
    _SDTM_DOMAIN_NAMES,
    NOT_SUBMITTED_BG,
    NOT_SUBMITTED_LABEL,
)


# =============================================================================
# COLOUR HELPERS
# =============================================================================

def _rgb_to_fitz(rgb: tuple[int, int, int]) -> tuple[float, float, float]:
    """Convert 0-255 RGB to 0-1 fitz colour."""
    return (rgb[0] / 255, rgb[1] / 255, rgb[2] / 255)


def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    """Convert RGB tuple to hex string for openpyxl."""
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


# =============================================================================
# EXPORT MANAGER
# =============================================================================

class ExportManager:
    """
    Manages all export operations for the annotation editor.

    Usage
    -----
    manager = ExportManager(store, registry, pdf_path)
    result  = manager.export_all(output_dir, session_id)
    """

    def __init__(
        self,
        store:    AnnotationStore,
        registry: DatasetColourRegistry,
        pdf_path: str | Path,
    ):
        self._store    = store
        self._registry = registry
        self._pdf_path = Path(pdf_path)

    # =========================================================================
    # MAIN EXPORT
    # =========================================================================

    def export_all(
        self,
        output_dir: str | Path,
        session_id: str = "session",
    ) -> dict[str, Path]:
        """
        Run all exports and return paths to output files.

        Parameters
        ----------
        output_dir : directory to write outputs
        session_id : used in output filenames

        Returns
        -------
        dict with keys:
          annotated_pdf, annotation_json,
          mapping_excel, report_excel
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        results = {}

        # 1 — Annotated PDF
        pdf_out = output_dir / f"CRF_Annotated_{session_id}_{ts}.pdf"
        self._export_annotated_pdf(pdf_out)
        results["annotated_pdf"] = pdf_out

        # 2 — Updated JSON
        json_out = output_dir / f"annotation_data_{session_id}_{ts}.json"
        self._export_json(json_out)
        results["annotation_json"] = json_out

        # 3 — Mapping Excel summary
        excel_out = output_dir / f"mapping_summary_{session_id}_{ts}.xlsx"
        self._export_mapping_excel(excel_out)
        results["mapping_excel"] = excel_out

        # 4 — Resolution report
        report_out = output_dir / f"resolution_report_{session_id}_{ts}.xlsx"
        self._export_resolution_report(report_out)
        results["report_excel"] = report_out

        return results

    # =========================================================================
    # 1 — ANNOTATED PDF
    # =========================================================================

    def _export_annotated_pdf(self, out_path: Path) -> None:
        """
        Produce annotated PDF using current store state.
        Mirrors crf_annotator.py logic exactly but uses
        the live store records including USER_CORRECTED
        and NOT_SUBMITTED statuses.
        """
        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent))
        from config import (
            FONT_NAME, FONT_BOLD,
            VAR_FONT_SIZE, DS_FONT_SIZE,
            CHIP_PAD_X, CHIP_PAD_Y, TEXT_BUFFER,
            DS_VERT_GAP, BORDER_WIDTH, LINE_MIN_RATIO,
            COL_BLACK, COL_BLUE, COL_RED,
            COL_LIGHT_RED, COL_WHITE,
            COL_DARK_GREY, COL_LIGHT_BLUE2,
            TOC_TITLE_SIZE, TOC_HEAD_SIZE, TOC_BODY_SIZE,
            dataset_label,
        )

        records  = self._store.to_records_list()
        by_page  = defaultdict(list)
        for rec in records:
            by_page[rec["page"]].append(rec)

        form_records  = [r for r in records if r.get("page_type") == "FORM"]
        resolved      = [r for r in form_records if r.get("sdtm_variable")]
        unresolved    = [
            r for r in form_records
            if not r.get("sdtm_variable")
            and r.get("status") != "NOT_SUBMITTED"
        ]
        not_submitted = [
            r for r in form_records
            if r.get("status") == "NOT_SUBMITTED"
        ]

        doc = fitz.open(str(self._pdf_path))

        for page_index in range(len(doc)):
            pnum         = page_index + 1
            page         = doc[page_index]
            page_records = by_page.get(pnum, [])

            if not page_records:
                continue

            page_type = page_records[0].get("page_type", "FORM")
            form_code = page_records[0].get("form_code", "")

            if page_type != "FORM":
                continue

            # Header zone
            header_y0, header_y1 = self._detect_header_zone(page)
            footer_y0             = self._get_footer_y(page)

            # Dataset chips
            ds_on_page = sorted(set(
                r["sdtm_dataset"].upper()
                for r in page_records
                if r.get("sdtm_dataset")
            ))
            if ds_on_page:
                self._draw_dataset_chips(
                    page, header_y0, header_y1,
                    form_code, ds_on_page,
                    FONT_NAME, FONT_BOLD,
                    DS_FONT_SIZE, CHIP_PAD_X, CHIP_PAD_Y,
                    TEXT_BUFFER, DS_VERT_GAP, BORDER_WIDTH,
                    COL_BLACK,
                )

            # Variable / unmapped / not-submitted boxes
            for rec in page_records:
                status    = rec.get("status", "")
                sdtm_ds   = rec.get("sdtm_dataset", "")
                sdtm_var  = rec.get("sdtm_variable", "")
                y0_pts    = rec.get("y0_pts", 0)
                y1_pts    = rec.get("y1_pts", 0)

                if y0_pts == 0 and y1_pts == 0:
                    continue
                if y0_pts >= footer_y0:
                    continue

                if status == "NOT_SUBMITTED":
                    self._draw_not_submitted_box(
                        page, y0_pts, y1_pts,
                        FONT_NAME, VAR_FONT_SIZE,
                        CHIP_PAD_X, CHIP_PAD_Y, TEXT_BUFFER,
                        BORDER_WIDTH,
                    )
                elif sdtm_ds and sdtm_var:
                    bg = self._get_colour(form_code, sdtm_ds)
                    self._draw_variable_box(
                        page, y0_pts, y1_pts,
                        f"{sdtm_ds}.{sdtm_var}", bg,
                        FONT_NAME, FONT_BOLD,
                        VAR_FONT_SIZE, CHIP_PAD_X, CHIP_PAD_Y,
                        TEXT_BUFFER, BORDER_WIDTH,
                        COL_BLUE,
                    )
                else:
                    self._draw_unresolved_box(
                        page, y0_pts, y1_pts,
                        FONT_NAME, VAR_FONT_SIZE,
                        CHIP_PAD_X, CHIP_PAD_Y, TEXT_BUFFER,
                        BORDER_WIDTH,
                    )

        # Build TOC
        self._build_toc(
            doc, unresolved, resolved,
            not_submitted, len(form_records),
            page_offset = 1,
            font_name   = FONT_NAME,
            font_bold   = FONT_BOLD,
            col_black   = COL_BLACK,
            col_blue    = COL_BLUE,
            col_red     = COL_RED,
            col_dark_grey = COL_DARK_GREY,
            col_light_red = COL_LIGHT_RED,
            col_light_blue2 = COL_LIGHT_BLUE2,
            toc_title_size  = TOC_TITLE_SIZE,
            toc_head_size   = TOC_HEAD_SIZE,
            toc_body_size   = TOC_BODY_SIZE,
        )

        out_path.parent.mkdir(parents=True, exist_ok=True)
        doc.save(str(out_path))
        doc.close()

    # ── Drawing helpers (mirrors crf_annotator.py) ────────────────────

    def _detect_header_zone(
        self, page: fitz.Page
    ) -> tuple[float, float]:
        page_w  = page.rect.width
        min_len = page_w * 0.4
        lines_y = []
        for path in page.get_drawings():
            for item in path.get("items", []):
                if item[0] == "l":
                    p1, p2 = item[1], item[2]
                    if (abs(p1.y - p2.y) < 2.0
                            and abs(p2.x - p1.x) >= min_len):
                        lines_y.append((p1.y + p2.y) / 2.0)
            r = path.get("rect")
            if r and r.height < 3.0 and r.width >= min_len:
                lines_y.append((r.y0 + r.y1) / 2.0)
        lines_y.sort()
        header_y1 = (
            lines_y[0] if lines_y
            else page.rect.height * 0.12
        )
        return 0.0, header_y1

    def _get_footer_y(self, page: fitz.Page) -> float:
        import re
        page_h      = page.rect.height
        study_id_re = re.compile(r'D\d+C\d+_\w+_V[\d.]+')
        footer_y    = None
        for w in page.get_text("words"):
            if study_id_re.search(w[4]) and w[1] > page_h * 0.5:
                if footer_y is None or w[1] < footer_y:
                    footer_y = w[1]
        return footer_y if footer_y is not None else page_h * 0.92

    def _get_colour(
        self, form_code: str, ds_code: str
    ) -> tuple[float, float, float]:
        rgb = self._registry.get_colour(ds_code, form_code)
        return _rgb_to_fitz(rgb)

    def _text_width(
        self, page: fitz.Page, text: str,
        fontsize: float, font_name: str,
        text_buffer: float,
    ) -> float:
        try:
            return (
                page.get_text_length(
                    text, fontname=font_name, fontsize=fontsize
                ) + text_buffer
            )
        except Exception:
            return 0.55 * fontsize * len(text) + text_buffer

    def _draw_box(
        self,
        page:          fitz.Page,
        x0: float, y0: float,
        x1: float, y1: float,
        bg_colour:     tuple,
        border_colour: tuple,
        text:          str,
        text_colour:   tuple,
        fontsize:      float,
        font_name:     str,
        border_width:  float,
        dashed:        bool = False,
    ) -> None:
        rect  = fitz.Rect(x0, y0, x1, y1)
        shape = page.new_shape()
        shape.draw_rect(rect)
        if dashed:
            shape.finish(
                color = border_colour,
                fill  = bg_colour,
                width = 1.2,
                dashes = "[3 3] 0",
            )
        else:
            shape.finish(
                color = border_colour,
                fill  = bg_colour,
                width = border_width,
            )
        shape.commit()

        if not text:
            return

        try:
            tw = page.get_text_length(
                text, fontname=font_name, fontsize=fontsize
            )
        except Exception:
            tw = 0.55 * fontsize * len(text)

        tx = x0 + (x1 - x0 - tw) / 2.0
        ty = y0 + (y1 - y0) * 0.65
        page.insert_text(
            fitz.Point(tx, ty), text,
            fontsize = fontsize,
            fontname = font_name,
            color    = text_colour,
        )

    def _draw_variable_box(
        self,
        page:         fitz.Page,
        y0_pts:       float,
        y1_pts:       float,
        sdtm_text:    str,
        bg_colour:    tuple,
        font_name:    str,
        font_bold:    str,
        fontsize:     float,
        chip_pad_x:   float,
        chip_pad_y:   float,
        text_buffer:  float,
        border_width: float,
        col_blue:     tuple,
    ) -> None:
        page_w  = page.rect.width
        comp_cy = (y0_pts + y1_pts) / 2.0
        tw      = self._text_width(
            page, sdtm_text, fontsize, font_name, text_buffer
        )
        box_w   = tw + chip_pad_x * 2
        box_h   = fontsize + chip_pad_y * 2

        x0 = max(page_w / 2.0 - box_w / 2.0, 4.0)
        x1 = min(page_w / 2.0 + box_w / 2.0, page_w - 4.0)
        y0 = max(comp_cy - box_h / 2.0, y0_pts + 1)
        y1 = min(comp_cy + box_h / 2.0, y1_pts - 1)

        if y1 <= y0 + 1 or x1 <= x0 + 1:
            return

        self._draw_box(
            page, x0, y0, x1, y1,
            bg_colour     = bg_colour,
            border_colour = col_blue,
            text          = sdtm_text,
            text_colour   = col_blue,
            fontsize      = fontsize,
            font_name     = font_name,
            border_width  = border_width,
        )

    def _draw_unresolved_box(
        self,
        page:         fitz.Page,
        y0_pts:       float,
        y1_pts:       float,
        font_name:    str,
        fontsize:     float,
        chip_pad_x:   float,
        chip_pad_y:   float,
        text_buffer:  float,
        border_width: float,
    ) -> None:
        page_w  = page.rect.width
        comp_cy = (y0_pts + y1_pts) / 2.0
        label   = "UNMAPPED"
        tw      = self._text_width(
            page, label, fontsize, font_name, text_buffer
        )
        box_w   = tw + chip_pad_x * 2
        box_h   = fontsize + chip_pad_y * 2

        x0 = max(page_w / 2.0 - box_w / 2.0, 4.0)
        x1 = min(page_w / 2.0 + box_w / 2.0, page_w - 4.0)
        y0 = max(comp_cy - box_h / 2.0, y0_pts + 1)
        y1 = min(comp_cy + box_h / 2.0, y1_pts - 1)

        if y1 <= y0 + 1 or x1 <= x0 + 1:
            return

        self._draw_box(
            page, x0, y0, x1, y1,
            bg_colour     = (1.0, 0.9, 0.9),
            border_colour = (0.8, 0.0, 0.0),
            text          = label,
            text_colour   = (0.8, 0.0, 0.0),
            fontsize      = fontsize,
            font_name     = font_name,
            border_width  = border_width,
            dashed        = True,
        )

    def _draw_not_submitted_box(
        self,
        page:         fitz.Page,
        y0_pts:       float,
        y1_pts:       float,
        font_name:    str,
        fontsize:     float,
        chip_pad_x:   float,
        chip_pad_y:   float,
        text_buffer:  float,
        border_width: float,
    ) -> None:
        page_w  = page.rect.width
        comp_cy = (y0_pts + y1_pts) / 2.0
        label   = "Not Submitted"
        tw      = self._text_width(
            page, label, fontsize, font_name, text_buffer
        )
        box_w   = tw + chip_pad_x * 2
        box_h   = fontsize + chip_pad_y * 2

        x0 = max(page_w / 2.0 - box_w / 2.0, 4.0)
        x1 = min(page_w / 2.0 + box_w / 2.0, page_w - 4.0)
        y0 = max(comp_cy - box_h / 2.0, y0_pts + 1)
        y1 = min(comp_cy + box_h / 2.0, y1_pts - 1)

        if y1 <= y0 + 1 or x1 <= x0 + 1:
            return

        # Grey bg, black border, black text
        self._draw_box(
            page, x0, y0, x1, y1,
            bg_colour     = (0.7, 0.7, 0.7),
            border_colour = (0.0, 0.0, 0.0),
            text          = label,
            text_colour   = (0.0, 0.0, 0.0),
            fontsize      = fontsize,
            font_name     = font_name,
            border_width  = border_width,
        )

    def _draw_dataset_chips(
        self,
        page:         fitz.Page,
        header_y0:    float,
        header_y1:    float,
        form_code:    str,
        ds_list:      list[str],
        font_name:    str,
        font_bold:    str,
        fontsize:     float,
        chip_pad_x:   float,
        chip_pad_y:   float,
        text_buffer:  float,
        ds_vert_gap:  float,
        border_width: float,
        col_black:    tuple,
    ) -> None:
        if not ds_list:
            return

        page_w      = page.rect.width
        start_x     = page_w / 2.0
        available_h = (header_y1 - header_y0) - 4.0

        fs = fontsize
        while fs >= 5.0:
            chip_h  = fs + chip_pad_y * 2
            total_h = (
                len(ds_list) * chip_h
                + (len(ds_list) - 1) * ds_vert_gap
            )
            if total_h <= available_h:
                break
            fs -= 0.5

        chip_h   = fs + chip_pad_y * 2
        cursor_y = header_y0 + 2.0

        for ds_code in ds_list:
            colour = self._get_colour(form_code, ds_code)
            name   = _SDTM_DOMAIN_NAMES.get(ds_code, ds_code)
            label  = (
                f"{name} ({ds_code})"
                if name != ds_code else ds_code
            )
            tw     = self._text_width(
                page, label, fs, font_name, text_buffer
            )
            chip_w = tw + chip_pad_x * 2

            x0 = start_x
            x1 = min(start_x + chip_w, page_w - 4.0)
            y0 = cursor_y
            y1 = cursor_y + chip_h

            if y1 > header_y1 - 1:
                break

            self._draw_box(
                page, x0, y0, x1, y1,
                bg_colour     = colour,
                border_colour = col_black,
                text          = label,
                text_colour   = col_black,
                fontsize      = fs,
                font_name     = font_name,
                border_width  = border_width,
            )
            cursor_y = y1 + ds_vert_gap

    # =========================================================================
    # TOC / CONTENTS PAGE
    # =========================================================================

    def _build_toc(
        self,
        doc:             fitz.Document,
        unresolved:      list[dict],
        resolved:        list[dict],
        not_submitted:   list[dict],
        total:           int,
        page_offset:     int,
        **style,
    ) -> None:
        """
        Build multi-page TOC with:
          - Summary stats
          - UNMAPPED section with GO TO links
          - NOT SUBMITTED section with GO TO links
          - RESOLVED summary by dataset
        """
        ref_page = doc[page_offset]
        pw       = ref_page.rect.width
        ph       = ref_page.rect.height
        ml       = 40.0
        mr       = pw - 40.0

        fn        = style["font_name"]
        fb        = style["font_bold"]
        c_black   = style["col_black"]
        c_blue    = style["col_blue"]
        c_red     = style["col_red"]
        c_dgrey   = style["col_dark_grey"]
        c_lred    = style["col_light_red"]
        c_lblue2  = style["col_light_blue2"]
        t_title   = style["toc_title_size"]
        t_head    = style["toc_head_size"]
        t_body    = style["toc_body_size"]

        NOT_SUB_GREY = (0.7, 0.7, 0.7)
        NOT_SUB_BLK  = (0.0, 0.0, 0.0)

        toc_pages: list[int] = []

        def new_toc_page() -> tuple[fitz.Page, float]:
            pos  = len(toc_pages)
            doc.insert_page(pos, width=pw, height=ph)
            toc_pages.append(pos)
            p = doc[pos]
            return p, 30.0

        def hline(
            p: fitz.Page, y: float,
            colour: tuple, width: float = 0.5,
        ) -> float:
            s = p.new_shape()
            s.draw_line(fitz.Point(ml, y), fitz.Point(mr, y))
            s.finish(color=colour, width=width)
            s.commit()
            return y + 10

        def check_overflow(
            p: fitz.Page, y: float, needed: float = 20.0
        ) -> tuple[fitz.Page, float]:
            if y + needed > ph - 30:
                np, ny = new_toc_page()
                np.insert_text(
                    fitz.Point(ml, ny),
                    "CRF ANNOTATION REVIEW  (continued)",
                    fontsize = t_head,
                    fontname = fb,
                    color    = c_dgrey,
                )
                ny += t_head + 10
                return np, ny
            return p, y

        def draw_table_header(
            p: fitz.Page, y: float
        ) -> float:
            col_page  = ml
            col_comp  = ml + 50
            col_field = ml + 155
            col_raw   = ml + 205
            col_prev  = ml + 300
            for hdr, cx in [
                ("Page",      col_page),
                ("Component", col_comp),
                ("Field #",   col_field),
                ("RAW Var",   col_raw),
                ("Preview",   col_prev),
            ]:
                p.insert_text(
                    fitz.Point(cx, y), hdr,
                    fontsize = 7.5,
                    fontname = fb,
                    color    = c_dgrey,
                )
            y += 10
            s = p.new_shape()
            s.draw_line(fitz.Point(ml, y), fitz.Point(mr, y))
            s.finish(color=c_dgrey, width=0.3)
            s.commit()
            return y + 6

        def draw_row_with_goto(
            p:        fitz.Page,
            y:        float,
            rec:      dict,
            row_bg:   tuple,
            col_page: float = ml,
            col_comp: float = ml + 50,
            col_field: float = ml + 155,
            col_raw:  float = ml + 205,
            col_prev: float = ml + 300,
            col_link: float = mr - 52,
        ) -> float:
            # Row background
            s = p.new_shape()
            s.draw_rect(fitz.Rect(ml - 2, y - 7, mr + 2, y + 5))
            s.finish(color=None, fill=row_bg, width=0)
            s.commit()

            raw_str  = rec.get("raw_variable") or "PENDING"
            prev_str = (rec.get("text_preview") or "")[:32]
            fn_str   = str(rec.get("field_number") or "?")

            for txt, cx in [
                (str(rec["page"]), col_page),
                (rec["component"], col_comp),
                (fn_str,           col_field),
                (raw_str,          col_raw),
                (prev_str,         col_prev),
            ]:
                p.insert_text(
                    fitz.Point(cx, y), txt,
                    fontsize = t_body,
                    fontname = fn,
                    color    = c_black,
                )

            # GO TO button
            lx0 = col_link
            lx1 = mr
            ly0 = y - 7
            ly1 = y + 5

            s2 = p.new_shape()
            s2.draw_rect(fitz.Rect(lx0, ly0, lx1, ly1))
            s2.finish(color=c_blue, fill=c_lblue2, width=0.6)
            s2.commit()

            p.insert_text(
                fitz.Point(lx0 + 3, y),
                "GO TO \u2192",
                fontsize = 7.5,
                fontname = fb,
                color    = c_blue,
            )

            target_page = rec["page"] - 1 + len(toc_pages)
            target_y    = rec.get("y0_pts", 0)

            p.insert_link({
                "kind": fitz.LINK_GOTO,
                "from": fitz.Rect(lx0, ly0, lx1, ly1),
                "page": target_page,
                "to":   fitz.Point(0, target_y),
                "zoom": 2.0,
            })

            return y + 14

        # ── Page 1 ────────────────────────────────────────────────────
        page, y = new_toc_page()

        # Title
        page.insert_text(
            fitz.Point(ml, y),
            "CRF ANNOTATION REVIEW",
            fontsize = t_title,
            fontname = fb,
            color    = c_black,
        )
        y += t_title + 4

        page.insert_text(
            fitz.Point(ml, y),
            f"Generated: {date.today().strftime('%d %b %Y')}",
            fontsize = t_body,
            fontname = fn,
            color    = c_dgrey,
        )
        y += t_body + 12

        y = hline(page, y, c_dgrey)

        # Stats
        res_pct  = len(resolved)      / total * 100 if total else 0
        ns_pct   = len(not_submitted) / total * 100 if total else 0
        unr_pct  = len(unresolved)    / total * 100 if total else 0

        for stat_line in [
            f"Total components annotated  :  {total}",
            f"Resolved (SDTM mapped)      :  {len(resolved)}  ({res_pct:.1f}%)",
            f"Not Submitted               :  {len(not_submitted)}  ({ns_pct:.1f}%)",
            f"Unresolved (needs review)   :  {len(unresolved)}  ({unr_pct:.1f}%)",
        ]:
            page.insert_text(
                fitz.Point(ml, y), stat_line,
                fontsize = t_body,
                fontname = fn,
                color    = c_black,
            )
            y += t_body + 5
        y += 8

        # ── UNRESOLVED SECTION ────────────────────────────────────────
        if unresolved:
            page, y = check_overflow(page, y, 40)
            y = hline(page, y, c_red, 0.8)
            page.insert_text(
                fitz.Point(ml, y),
                "UNRESOLVED COMPONENTS  \u2014  click to navigate",
                fontsize = t_head,
                fontname = fb,
                color    = c_red,
            )
            y += t_head + 8

            y = draw_table_header(page, y)

            for rec in unresolved:
                page, y = check_overflow(page, y, 14)
                if y < 80:
                    y = draw_table_header(page, y)
                y = draw_row_with_goto(page, y, rec, c_lred)

            y += 10

        # ── NOT SUBMITTED SECTION ─────────────────────────────────────
        if not_submitted:
            page, y = check_overflow(page, y, 40)
            y = hline(page, y, NOT_SUB_BLK, 0.8)
            page.insert_text(
                fitz.Point(ml, y),
                "NOT SUBMITTED COMPONENTS  \u2014  click to navigate",
                fontsize = t_head,
                fontname = fb,
                color    = NOT_SUB_BLK,
            )
            y += t_head + 8

            y = draw_table_header(page, y)

            for rec in not_submitted:
                page, y = check_overflow(page, y, 14)
                if y < 80:
                    y = draw_table_header(page, y)
                y = draw_row_with_goto(
                    page, y, rec, NOT_SUB_GREY
                )

            y += 10

        # ── RESOLVED SUMMARY ──────────────────────────────────────────
        page, y = check_overflow(page, y, 40)
        y = hline(page, y, c_blue, 0.8)
        page.insert_text(
            fitz.Point(ml, y),
            "RESOLVED COMPONENTS SUMMARY",
            fontsize = t_head,
            fontname = fb,
            color    = c_blue,
        )
        y += t_head + 8

        by_ds: dict[str, list[dict]] = defaultdict(list)
        for rec in resolved:
            by_ds[rec.get("sdtm_dataset", "UNKNOWN")].append(rec)

        for ds, recs in sorted(by_ds.items()):
            page, y = check_overflow(page, y, 20)
            name    = _SDTM_DOMAIN_NAMES.get(ds, ds)
            page.insert_text(
                fitz.Point(ml, y),
                f"{name} ({ds})  —  {len(recs)} variables",
                fontsize = t_body,
                fontname = fb,
                color    = c_dgrey,
            )
            y += t_body + 4

            col_w   = (mr - ml) / 3.0
            col_idx = 0
            row_y   = y

            for rec in recs:
                page, row_y = check_overflow(page, row_y, 9)
                entry = (
                    f"  {rec.get('raw_variable','?')} \u2192 "
                    f"{ds}.{rec.get('sdtm_variable','?')}"
                )
                cx = ml + col_idx * col_w
                page.insert_text(
                    fitz.Point(cx, row_y), entry,
                    fontsize = 7.0,
                    fontname = fn,
                    color    = c_dgrey,
                )
                col_idx += 1
                if col_idx >= 3:
                    col_idx = 0
                    row_y  += 9

            y = row_y + 12

    # =========================================================================
    # 2 — EXPORT JSON
    # =========================================================================

    def _export_json(self, out_path: Path) -> None:
        """Export current store state as annotation_data.json."""
        records = self._store.to_records_list()
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)

    # =========================================================================
    # 3 — MAPPING EXCEL SUMMARY
    # =========================================================================

    def _export_mapping_excel(self, out_path: Path) -> None:
        """
        Export RAW → SDTM mapping summary to Excel.
        One row per annotation, colour-coded by status.
        """
        records = self._store.to_records_list()

        rows = []
        for rec in records:
            if rec.get("page_type") != "FORM":
                continue
            rows.append({
                "Page":          rec.get("page", ""),
                "Form":          rec.get("form_code", ""),
                "Component":     rec.get("component", ""),
                "Field #":       rec.get("field_number", ""),
                "RAW Variable":  rec.get("raw_variable", ""),
                "SDTM Dataset":  rec.get("sdtm_dataset", ""),
                "SDTM Variable": rec.get("sdtm_variable", ""),
                "SDTM Label":    rec.get("sdtm_label", ""),
                "Status":        rec.get("status", ""),
                "Preview":       (rec.get("text_preview") or "")[:60],
            })

        df = pd.DataFrame(rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Annotation Mappings"

        # Header row
        headers = list(df.columns)
        header_fill = PatternFill(
            "solid", fgColor="003366"
        )
        header_font = Font(
            bold=True, color="FFFFFF", size=10
        )
        for col_idx, header in enumerate(headers, 1):
            cell            = ws.cell(row=1, column=col_idx)
            cell.value      = header
            cell.fill       = header_fill
            cell.font       = header_font
            cell.alignment  = Alignment(horizontal="center")

        # Status colour map
        status_fills = {
            "RESOLVED":       PatternFill("solid", fgColor="D6EAF8"),
            "USER_CORRECTED": PatternFill("solid", fgColor="D5F5E3"),
            "UNMAPPED":       PatternFill("solid", fgColor="FADBD8"),
            "NOT_SUBMITTED":  PatternFill("solid", fgColor="D5D8DC"),
        }

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            status = row.get("Status", "")
            fill   = status_fills.get(status)
            for col_idx, header in enumerate(headers, 1):
                cell           = ws.cell(
                    row=row_idx, column=col_idx
                )
                cell.value     = row.get(header, "")
                cell.alignment = Alignment(
                    horizontal="left", wrap_text=False
                )
                if fill:
                    cell.fill = fill

        # Auto column width
        for col in ws.columns:
            max_len = max(
                len(str(cell.value or ""))
                for cell in col
            )
            ws.column_dimensions[
                col[0].column_letter
            ].width = min(max_len + 4, 40)

        # Freeze header
        ws.freeze_panes = "A2"

        wb.save(str(out_path))

    # =========================================================================
    # 4 — RESOLUTION REPORT
    # =========================================================================

    def _export_resolution_report(self, out_path: Path) -> None:
        """
        Export resolution report Excel with:
          - Summary sheet
          - Per-domain breakdown
          - Unmapped list
          - Not submitted list
        """
        records      = self._store.to_records_list()
        form_records = [
            r for r in records
            if r.get("page_type") == "FORM"
        ]

        resolved      = [r for r in form_records if r.get("sdtm_variable")]
        unresolved    = [
            r for r in form_records
            if not r.get("sdtm_variable")
            and r.get("status") != "NOT_SUBMITTED"
        ]
        not_submitted = [
            r for r in form_records
            if r.get("status") == "NOT_SUBMITTED"
        ]
        user_corrected = [
            r for r in form_records
            if r.get("status") == "USER_CORRECTED"
        ]

        total    = len(form_records)
        res_pct  = len(resolved) / total * 100 if total else 0

        wb = openpyxl.Workbook()

        # ── Summary sheet ─────────────────────────────────────────────
        ws_sum        = wb.active
        ws_sum.title  = "Summary"

        title_font  = Font(bold=True, size=14, color="003366")
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="003366")

        ws_sum["A1"] = "CRF Annotation Resolution Report"
        ws_sum["A1"].font = title_font
        ws_sum["A2"] = f"Generated: {date.today().strftime('%d %b %Y')}"
        ws_sum["A2"].font = Font(italic=True, color="666666")

        ws_sum.append([])
        ws_sum.append(["Metric", "Count", "Percentage"])
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = header_fill
            cell.font = header_font

        summary_rows = [
            ("Total Components",    total,               "100%"),
            ("Resolved",            len(resolved),       f"{res_pct:.1f}%"),
            ("User Corrected",      len(user_corrected), f"{len(user_corrected)/total*100:.1f}%" if total else "0%"),
            ("Unresolved",          len(unresolved),     f"{len(unresolved)/total*100:.1f}%" if total else "0%"),
            ("Not Submitted",       len(not_submitted),  f"{len(not_submitted)/total*100:.1f}%" if total else "0%"),
        ]

        fill_map = [
            None,
            PatternFill("solid", fgColor="D6EAF8"),
            PatternFill("solid", fgColor="D5F5E3"),
            PatternFill("solid", fgColor="FADBD8"),
            PatternFill("solid", fgColor="D5D8DC"),
        ]

        for i, (metric, count, pct) in enumerate(summary_rows):
            ws_sum.append([metric, count, pct])
            if fill_map[i]:
                for cell in ws_sum[ws_sum.max_row]:
                    cell.fill = fill_map[i]

        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 12
        ws_sum.column_dimensions["C"].width = 14

        # ── Per-domain sheet ──────────────────────────────────────────
        ws_ds        = wb.create_sheet("By Domain")
        ws_ds.append(["Domain", "Full Name", "Count", "Variables"])
        for cell in ws_ds[1]:
            cell.fill = header_fill
            cell.font = header_font

        by_ds: dict[str, list] = defaultdict(list)
        for rec in resolved:
            by_ds[rec.get("sdtm_dataset", "?")].append(
                rec.get("sdtm_variable", "?")
            )

        for ds, vars_list in sorted(by_ds.items()):
            ws_ds.append([
                ds,
                _SDTM_DOMAIN_NAMES.get(ds, ds),
                len(vars_list),
                ", ".join(sorted(set(vars_list))),
            ])

        for col in ws_ds.columns:
            ws_ds.column_dimensions[
                col[0].column_letter
            ].width = min(
                max(len(str(c.value or "")) for c in col) + 4,
                60
            )

        # ── Unmapped sheet ────────────────────────────────────────────
        ws_unr        = wb.create_sheet("Unresolved")
        ws_unr.append([
            "Page", "Form", "Component",
            "Field #", "RAW Variable", "Preview"
        ])
        for cell in ws_unr[1]:
            cell.fill = header_fill
            cell.font = header_font

        for rec in unresolved:
            ws_unr.append([
                rec.get("page", ""),
                rec.get("form_code", ""),
                rec.get("component", ""),
                rec.get("field_number", ""),
                rec.get("raw_variable", ""),
                (rec.get("text_preview") or "")[:60],
            ])

        # ── Not submitted sheet ───────────────────────────────────────
        ws_ns        = wb.create_sheet("Not Submitted")
        ws_ns.append([
            "Page", "Form", "Component",
            "Field #", "RAW Variable", "Preview"
        ])
        for cell in ws_ns[1]:
            cell.fill = header_fill
            cell.font = header_font

        for rec in not_submitted:
            ws_ns.append([
                rec.get("page", ""),
                rec.get("form_code", ""),
                rec.get("component", ""),
                rec.get("field_number", ""),
                rec.get("raw_variable", ""),
                (rec.get("text_preview") or "")[:60],
            ])

        wb.save(str(out_path))

    # =========================================================================
    # 5 — FEED BACK TO MASTER MAPPING EXCEL
    # =========================================================================

    def feed_back_to_mapping_excel(
        self,
        mapping_excel_path: str | Path,
        backup:             bool = True,
    ) -> int:
        """
        Feed USER_CORRECTED mappings back into the master
        RAW_SDTM_Mappings.xlsx so future pipeline runs
        auto-resolve them.

        Parameters
        ----------
        mapping_excel_path : path to RAW_SDTM_Mappings.xlsx
        backup             : if True, backs up the file first

        Returns
        -------
        int — number of new mappings added
        """
        mapping_excel_path = Path(mapping_excel_path)

        if backup:
            ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak_path = mapping_excel_path.with_suffix(
                f".bak_{ts}.xlsx"
            )
            shutil.copy2(mapping_excel_path, bak_path)

        # Get all USER_CORRECTED records
        records        = self._store.to_records_list()
        new_mappings   = [
            r for r in records
            if r.get("status") == "USER_CORRECTED"
            and r.get("raw_variable")
            and r.get("sdtm_variable")
        ]

        if not new_mappings:
            return 0

        # Load existing workbook
        wb = openpyxl.load_workbook(str(mapping_excel_path))

        RAW_SDTM_SHEET = "RAW-SDTM Mappings"

        if RAW_SDTM_SHEET not in wb.sheetnames:
            # Create sheet if missing
            ws = wb.create_sheet(RAW_SDTM_SHEET)
            ws.append([
                "Source Dataset", "Source Variable",
                "Source Label", "Source Type",
                "Source Format", "Source Length",
                "Source Codelist", "Source Notes",
                "SDTM Dataset", "SDTM Variable", "SDTM Label",
            ])
        else:
            ws = wb[RAW_SDTM_SHEET]

        # Build set of existing (src_dataset, src_var) pairs
        existing = set()
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and row[1]:
                existing.add(
                    (str(row[1]).strip().upper(),)
                )

        added = 0
        for rec in new_mappings:
            raw_var = rec["raw_variable"].strip().upper()
            if (raw_var,) in existing:
                continue

            ws.append([
                rec.get("form_code", ""),
                raw_var,
                rec.get("text_preview", "")[:40],
                "",  # source type
                "",  # source format
                "",  # source length
                "",  # source codelist
                f"Added by editor {date.today()}",
                rec.get("sdtm_dataset", ""),
                rec.get("sdtm_variable", ""),
                rec.get("sdtm_label", ""),
            ])
            existing.add((raw_var,))
            added += 1

        wb.save(str(mapping_excel_path))
        return added